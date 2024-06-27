from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Alignment, Font
import os
from datetime import datetime

def obter_dados(loja, url, xpath):
    options = Options()
    options.headless = True
    service = Service()

    driver = webdriver.Chrome(service=service, options=options)
    driver.get(url)
    
    try:
        elemento = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        preco = elemento.text.strip()
        if loja == 'Loja oficial Samsung':
            preco = preco[2:7]
        if loja == 'Casas Bahia':
            preco = preco[7:12]
    except Exception as e:
        preco = "não encontrado"
    finally:
        driver.quit()
    
    return {"LOJA": loja, "PRECO": preco}

def main():
    resultado = [
        obter_dados('Amazon', 'https://www.amazon.com.br/dp/B0CYNG67P9/', '//*[@id="corePrice_feature_div"]/div/div/span[1]/span[2]/span[2]'),
        obter_dados('Fast Shop', 'https://www.fastshop.com.br/web/p/d/3005647532_PRD/samsung-smart-43-crystal-uhd-4k-43du8000-2024-painel-dynamic-crystal-color-alexa-built-in', '//*[@id="auto_pdp_price_content"]/div[1]/app-price-payments/div[1]/span[1]/span[2]'),
        obter_dados('Loja oficial Samsung', 'https://www.samsung.com/br/tvs/uhd-4k-tv/du8000-43-inch-crystal-uhd-4k-tizen-os-smart-tv-un43du8000gxzd/', '//*[@id="anchorContainer"]/div[2]/div[2]/div[1]'),
        obter_dados('Casas Bahia', 'https://www.casasbahia.com.br/samsung-smart-tv-43-quot-crystal-uhd-4k-43du8000-2024-painel-dynamic-crystal-color-alexa-built-in/p/1566004815?utm_source=zoom&utm_medium=comparadorpreco&utm_campaign=1ad9f36014c24b0699cd3f31221bf2a6', '//*[@id="product-price"]/span[1]')
    ]

    criar_planilha(resultado)
    
def criar_planilha(resultado):
    
    filename = "resultados.xlsx"
    
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
    
    data_hora_atual = datetime.now().strftime('%d-%m-%Y, %Hh%Mmin%Sseg')
    ws = wb.create_sheet(title=f"{data_hora_atual}") #define o título da planilha
    
    ws.merge_cells('A1:B1') #mescla células para definir o conteúdo do titulo da tabela; define alinhamento e tamanho da fonte também
    cell = ws['A1']
    cell.value = f"{data_hora_atual}"
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(size=18, bold=True)

    ws.append(["LOJA", "PREÇO"]) #define as colunas e seus valores
    for item in resultado:
        ws.append([item["LOJA"], item["PRECO"]])
    
    for col_idx in range(1, ws.max_column + 1): #define as dimensões das células
        max_comprimento = 0
        for cell in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for c in cell:
                try:
                    if len(str(c.value)) > max_comprimento:
                        max_comprimento = len(str(c.value))
                except:
                    pass
        largura_ajustada = (max_comprimento + 2) * 1.2
        col_letra = openpyxl.utils.cell.get_column_letter(col_idx)
        ws.column_dimensions[col_letra].width = largura_ajustada
    
    #formata os valores da coluna preço como real sem centavos
    preco_col_idx = 2
    for cell in ws.iter_cols(min_col=preco_col_idx, max_col=preco_col_idx, min_row=3, max_row=ws.max_row):
        for c in cell:
            try:
                valor = int(c.value.replace('.', '').replace(',', '').replace('R$', '').strip())
                c.value = valor
                c.number_format = 'R$ #,##0'
            except ValueError:
                pass

    wb.save(filename)
    os.system(f'start {filename}')

if __name__ == "__main__":
    main()
